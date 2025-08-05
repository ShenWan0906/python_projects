import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from deep_translator import GoogleTranslator
from datetime import datetime
import threading

# 自动调整excel列宽
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

# 创建唯一输出文件夹
def create_unique_output_dir(base_dir: str, prefix: str) -> str:
    date_str = datetime.now().strftime("%Y%m%d")
    base_name = f"{prefix}_{date_str}"
    output_path = os.path.join(base_dir, base_name)

    count = 1
    while os.path.exists(output_path):
        output_path = os.path.join(base_dir, f"{base_name}_{count}")
        count += 1

    os.makedirs(output_path)
    return output_path

def translate_file(filepath, update_progress_callback, on_complete_callback):
    try:
        prefix = "contractors.manage."
        translator = GoogleTranslator(source='en', target='ar')

        wb = load_workbook(filepath)
        ws = wb.active

        en_dict = {}
        ar_dict = {}
        usage_excel_data = []

        rows = [row for row in ws.iter_rows(min_row=1, max_col=1) if row[0].value and row[0].value.strip()]
        total = len(rows)

        # 用当前时间戳到毫秒，保证前半段唯一
        timestamp_str = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]  # 精确到毫秒
        seq = 1  # 当次文件内序号

        translation_cache = {}

        for i, row in enumerate(rows):
            en_text = row[0].value.strip()
            key = f"{prefix}{timestamp_str}{seq:03d}"  # 后面seq补3位，不够补0

            if en_text in translation_cache:
                ar_text = translation_cache[en_text]
            else:
                try:
                    ar_text = translator.translate(en_text)
                    translation_cache[en_text] = ar_text
                except Exception as e:
                    ar_text = f"ERROR: {e}"

            en_dict[key] = en_text
            ar_dict[key] = ar_text
            usage_excel_data.append((en_text, f"$t('{key}')"))
            seq += 1

            update_progress_callback(i + 1, total)

        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        output_dir = create_unique_output_dir(desktop_path, "i18n_output")

        with open(os.path.join(output_dir, "en_output.js"), "w", encoding="utf-8") as f:
            f.write("export default ")
            json.dump(en_dict, f, ensure_ascii=False, indent=2)

        with open(os.path.join(output_dir, "ar_output.js"), "w", encoding="utf-8") as f:
            f.write("export default ")
            json.dump(ar_dict, f, ensure_ascii=False, indent=2)

        usage_wb = Workbook()
        usage_ws = usage_wb.active
        usage_ws.title = "Usage Guide"
        usage_ws.append(["English Text", "Vue Usage Key"])
        for en, usage in usage_excel_data:
            usage_ws.append([en, usage])

        auto_adjust_column_width(usage_ws)

        usage_wb.save(os.path.join(output_dir, "t_usage_snippet.xlsx"))

        on_complete_callback(success=True, output_dir=output_dir)
    except Exception as e:
        on_complete_callback(success=False, error=str(e))

def launch_app():
    root = tk.Tk()
    root.title("📄 i18n 英文转阿拉伯文工具")
    root.geometry("420x260")
    root.resizable(False, False)

    title_label = tk.Label(root, text="Excel 多语言翻译工具", font=("Arial", 16))
    title_label.pack(pady=20)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100, length=300)
    progress_bar.pack(pady=10)

    progress_label = tk.Label(root, text="请点击下方按钮开始翻译", font=("Arial", 10))
    progress_label.pack()

    def update_progress(current, total):
        percent = (current / total) * 100
        progress_var.set(percent)
        progress_label.config(text=f"翻译中：{current} / {total}")

    def on_complete(success, output_dir=None, error=None):
        if success:
            messagebox.showinfo("✅ 完成", f"翻译完成，文件保存在：\n{output_dir}")
            os.startfile(output_dir)
        else:
            messagebox.showerror("❌ 错误", f"翻译失败：{error}")
        progress_var.set(0)
        progress_label.config(text="请选择新的 Excel 文件继续")

    def choose_file():
        file_path = filedialog.askopenfilename(
            title="请选择 Excel 文件",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            threading.Thread(
                target=translate_file,
                args=(file_path, update_progress, on_complete),
                daemon=True
            ).start()

    select_btn = tk.Button(root, text="选择 Excel 文件进行翻译", font=("Arial", 12), command=choose_file)
    select_btn.pack(pady=15)

    root.mainloop()

if __name__ == "__main__":
    launch_app()
